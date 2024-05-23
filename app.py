#importar todas as bibliotecas utilizada
import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl
import openpyxl.workbook
import xlrd
import pathlib
from openpyxl import Workbook
from datetime import datetime
import pandas as pd
import pandas as pd
import tkinter as tk
from tkinter import ttk
import base64
from tkcalendar import Calendar, DateEntry
import datetime


# aparencias do sistema 
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.aparencia()
        self.todo_sistema()
        self.lerInf()
        
        
        
    
    
    
    def layout_config(self):
        self.title("Sistema de cobrança automatico")
        self.geometry("1024x768")    
    
    
    def aparencia(self):
        self.lb_apm = ctk.CTkLabel(self, text="tema", bg_color ='transparent', text_color=["#000", "#fff"]).place(x=10, y=690)
        self.opt_apm = ctk.CTkOptionMenu(self, values=["Ligth", "Dark", "System"], command=self.change_apm).place(x=10, y=720)
    
   
    
    
    
    def todo_sistema(self):
        
        freme = ctk.CTkFrame(self, width= 1024, height= 60, corner_radius=0, bg_color="teal", fg_color="teal")
        freme.place(x=0, y=10)
        title = ctk.CTkLabel(freme, text="Sistema Palmed", font=("Century Gothic bold", 24), text_color="#fff").place(x=400, y=25)
        
        span = ctk.CTkLabel(self, text="Por favor prencha o formulario", font=("Century Gothic bold", 18), text_color=["#000", "#fff"]).place(x=50, y=90)
        
        lista = pathlib.Path('teste.xlsx')
        
        if lista.exists():
            pass
        else:
            lista = openpyxl.Workbook()
            folha = lista.active
            folha['A1'] = "Descrição"
            folha['B1'] = "Data Inicial"
            folha['C1'] = "Data Final"
            folha['D1'] = "Uf"
            folha['E1'] = "Origem"
            folha['F1'] = "Tipo Restrição"
            folha['G1'] = "Promoção"
            folha['H1'] = "Execução"
            folha['I1'] = "Vigencia"
            folha['J1'] = "Truno"
            lista.save('teste.xlsx')
        #botoes de salvar e limpar
        def submit():
            
            
            #pegando os dados entrys
            descricao = descricao_value.get()  
            origem = origem_value.get() 
            promocao = promocao_value.get()  
            vigencia = vingencia_value.get()
            uf = comobox_uf.get()
            truno = comobox_turno.get()
            tipo_restricao = comobox_tipo_restricao.get()
            execucao = comobox_execucao.get()
            
            try:
                inicio_str = inicio_value.get()
                date_obj = datetime.datetime.strptime(inicio_str, "%d/%m/%Y")
                inicio = date_obj.strftime("%d/%m/%Y")
                final_str = inicio_value.get()
                date_obj = datetime.datetime.strptime(final_str, "%d/%m/%Y")
                final = date_obj.strftime("%d/%m/%Y")
            except:
                messagebox.showerror("Erro", "ERRO!\n Por favor insira a data corretamente!")
            
                
            if(descricao == "" or inicio == "" or final == "" or origem == "" or promocao == "" or vigencia == "" or uf == "" or truno == "" or tipo_restricao == "" or execucao == ""):
                messagebox.showerror("Erro", "ERRO!\n Por favor precher todos os campos")
            else:
            
                lista = openpyxl.load_workbook('teste.xlsx')
                folha = lista.active
                folha.cell(column = 1, row=folha.max_row+1, value=descricao)
                folha.cell(column = 2, row=folha.max_row, value=inicio)
                folha.cell(column = 3, row=folha.max_row, value=final)
                folha.cell(column = 4, row=folha.max_row, value=uf)
                folha.cell(column = 5, row=folha.max_row, value=origem)
                folha.cell(column = 6, row=folha.max_row, value=tipo_restricao)
                folha.cell(column = 7, row=folha.max_row, value=promocao)
                folha.cell(column = 8, row=folha.max_row, value=execucao)
                folha.cell(column = 9, row=folha.max_row, value=vigencia)
                folha.cell(column = 10, row=folha.max_row, value=truno)  
                
                lista.save(r'teste.xlsx')
                self.lerInf()
                messagebox.showinfo("Sistema", "Dados salvos com sucesso!")
                
                
                
        
        def clear():
            descricao_value.set("")
            inicio_value.set("") 
            final_value.set("")  
            origem_value.set("") 
            promocao_value.set("")  
            vingencia_value.set("")



        #Combobox (criar caixinha) para marcar
        
        comobox_uf = ctk.CTkComboBox(self, values=["AC", "AL", "AP", "AM", "BA", "CE", "DF", "ES", "GO", "MA", 
                                                    "MT", "MS", "MG", "PA", "PB", "PR", "PE", "PI", "RJ", "RN", 
                                                    "RS", "RO", "RR", "SC", "SP", "SE", "TO"
                                                ], font=("Century Gohtic bold", 14))
        comobox_uf.set("AC")
    
        comobox_tipo_restricao = ctk.CTkComboBox(self, values=["R", "A"], font=("Century Gohtic bold", 14))
        comobox_tipo_restricao.set("R")
        
        comobox_turno = ctk.CTkComboBox(self, values=["Matutino", "Vespertino"], font=("Century Gohtic bold", 14))
        comobox_turno.set("Matutino")
        
        comobox_execucao = ctk.CTkComboBox(self, values=[    "DOM", "SEG", "TER",
                                                            "QUA", "QUI", "SEX",
                                                            "SAB"
                                                        ], font=("Century Gohtic bold", 14))
        comobox_execucao.set("SEG")
        
        #Entrada de obs
        obs_entry = ctk.CTkTextbox(self, width=500, height=150, font=("arial", 18), border_color="#aaa", border_width=2, fg_color="transparent")        
        
        
        #definindo os tipos das variaveis
        descricao_value = StringVar()
        inicio_value = StringVar()        
        final_value = StringVar()
        origem_value = StringVar()
        promocao_value = StringVar()
        vingencia_value = IntVar()
        #ENTRYS (CAIXAS PARA PREENCHER)
        entry_descricao = ctk.CTkEntry(self, width=300, textvariable=descricao_value, font=("Century Gohtic bold", 16), fg_color="transparent")
        
        entry_dt_inicio = ctk.CTkEntry(self, width=300, textvariable=inicio_value, font=("Century Gohtic bold", 16), fg_color="transparent")
        
        entry_dt_fim = ctk.CTkEntry(self, width=300, textvariable=final_value,font=("Century Gohtic bold", 16), fg_color="transparent")
        
        entry_origem = ctk.CTkEntry(self, width=300, textvariable=origem_value,font=("Century Gohtic bold", 16), fg_color="transparent")
        
        entry_id_promocao = ctk.CTkEntry(self, width=300, textvariable=promocao_value,font=("Century Gohtic bold", 16), fg_color="transparent")
        
        entry_vigencia = ctk.CTkEntry(self, width=300, textvariable=vingencia_value,font=("Century Gohtic bold", 16), fg_color="transparent")

        
        
            
        
        
        
        
        #todos os campos 
        lb_DESCRICAO = ctk.CTkLabel(self, text="Descrição", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        
        lb_DT_INICIO = ctk.CTkLabel(self, text="Data Inicial: ", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        
        lb_DT_FIM = ctk.CTkLabel(self, text="Data Final: ", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        
        lb_UF = ctk.CTkLabel(self, text="UF-ESTADO", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        
        lb_ORIGEM = ctk.CTkLabel(self, text="Origem", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        
        lb_TIPO_RESTRICAO = ctk.CTkLabel(self, text="Tipo Restriçao", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        
        lb_ID_PROMOCAO = ctk.CTkLabel(self, text="Id Promoção", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        
        lb_EXECUCAO = ctk.CTkLabel(self, text="Execução", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        
        lb_VIGENCIA = ctk.CTkLabel(self, text="Vigencia", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        
        lb_TURNO = ctk.CTkLabel(self, text="Turno(Matutino/Vespertino)", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        
    
        btn_submit = ctk.CTkButton(self, text="Enviar Dados".upper(), command=submit, fg_color="#151", hover_color="#131").place(x=320, y=720)
        
        btn_submit = ctk.CTkButton(self, text="Limpar Dados".upper(), command=clear, fg_color="#555", hover_color="#333").place(x=490, y=720)
        
        bt_atualiza_tabela = ctk.CTkButton(self, text='Atualizar'.upper(), command=self.lerInf, fg_color="#151", hover_color="#131").place(x=660, y=720)
        
        
    #posicoes nas janelas
        lb_DESCRICAO.place(x=50, y=120)
        entry_descricao.place(x=50, y=150)
        
        lb_DT_INICIO.place(x=380, y=120)
        entry_dt_inicio.place(x=380, y=150)
        
        lb_DT_FIM.place(x=700, y=120)
        entry_dt_fim.place(x=700, y=150)
        
        lb_ORIGEM.place(x=700, y=190)
        entry_origem.place(x=700, y=220)
        
        lb_ID_PROMOCAO.place(x=50, y=190)
        entry_id_promocao.place(x=50, y=220)
        
        lb_VIGENCIA.place(x=380, y=190)
        entry_vigencia.place(x=380, y=220)
        
        lb_EXECUCAO.place(x=50, y=260)
        comobox_execucao.place(x=50, y=300)
        
        lb_UF.place(x=300, y=260)
        comobox_uf.place(x=300, y=300)
        
        lb_TURNO.place(x=600, y=260)
        comobox_turno.place(x=600, y=300)
        
        lb_TIPO_RESTRICAO.place(x=860, y=260)
        comobox_tipo_restricao.place(x=860, y=300)
    
    
    def change_apm(self, nova_aparencia):
        ctk.set_appearance_mode(nova_aparencia)
        
    
    
    
    def lerInf(self):
        ler = pd.read_excel('teste.xlsx')
        table_frame = tk.Frame(self)
        table_frame.pack(padx=10, pady=10)
        columns = ["Descrição", "Vigencia", "Uf", "Data Inicial", "Data Final", "Origem", "Tipo Restrição", "Promoção", "Execução", "Truno"]
        table = ttk.Treeview(table_frame, columns=columns, show="headings")
        table.heading("#0", text="") 
        table.column("Descrição", width=100, minwidth=50, stretch=NO)
        table.column("Vigencia", width=100, minwidth=50, stretch=NO)
        table.column("Uf", width=100, minwidth=50, stretch=NO)
        table.column("Data Inicial", width=100, minwidth=50, stretch=NO)
        table.column("Data Final", width=100, minwidth=50, stretch=NO)
        table.column("Origem", width=100, minwidth=50, stretch=NO)
        table.column("Tipo Restrição", width=100, minwidth=50, stretch=NO)
        table.column("Promoção", width=100, minwidth=50, stretch=NO)
        table.column("Execução", width=100, minwidth=50, stretch=NO)
        table.column("Truno", width=100, minwidth=50, stretch=NO)
        table_frame.place(x=11, y=400)
        for i, col in enumerate(columns):
            table.heading(col, text=col)
        table.pack(fill=tk.BOTH, expand=True)
        for index, row in ler.iterrows():
            values = [row[col] for col in columns]
            table.insert("", index, values=values)
        
    
    
        
if __name__=="__main__":
    app = App()
    app.mainloop()
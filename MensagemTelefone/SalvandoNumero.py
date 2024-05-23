import tkinter as tk
from tkinter import ttk
import openpyxl

def salvar_dados():
  nome = entry_nome.get()
  telefone = entry_telefone.get()

  # Verificar se os campos estão preenchidos
  if not nome or not telefone:
    mensagem_erro.config(text="Preencha todos os campos!")
    return

  # Abrir ou criar arquivo Excel
  try:
    arquivo = openpyxl.load_workbook("Numeros.xlsx")
  except FileNotFoundError:
    arquivo = openpyxl.Workbook()

  # Acessar planilha ativa ou criar nova
  planilha = arquivo.active
  if not planilha["A1"].value:
    planilha.append(["Nome", "Telefone"])

  # Inserir dados na próxima linha disponível
  ultima_linha = planilha.max_row + 1
  planilha.cell(row=ultima_linha, column=1).value = nome
  planilha.cell(row=ultima_linha, column=2).value = telefone

  # Salvar arquivo Excel
  arquivo.save(r"Numeros.xlsx")

  # Limpar campos de entrada e exibir mensagem de sucesso
  entry_nome.delete(0, tk.END)
  entry_telefone.delete(0, tk.END)
  mensagem_erro.config(text="Dados salvos com sucesso!")

# Criar interface gráfica
janela = tk.Tk()
janela.title("Coleta de Dados")

# Rótulos e campos de entrada
label_nome = tk.Label(janela, text="Nome:")
label_nome.grid(row=0, column=0, padx=5, pady=5)
entry_nome = tk.Entry(janela)
entry_nome.grid(row=0, column=1, padx=5, pady=5)

label_telefone = tk.Label(janela, text="Telefone:")
label_telefone.grid(row=1, column=0, padx=5, pady=5)
entry_telefone = tk.Entry(janela)
entry_telefone.grid(row=1, column=1, padx=5, pady=5)

# Botão para salvar dados
botao_salvar = tk.Button(janela, text="Salvar", command=salvar_dados)
botao_salvar.grid(row=2, column=0, columnspan=2, padx=5, pady=5)

# Mensagem de erro
mensagem_erro = tk.Label(janela, text="", fg="red")
mensagem_erro.grid(row=3, column=0, columnspan=2, padx=5, pady=5)

# Executar interface gráfica
janela.mainloop()